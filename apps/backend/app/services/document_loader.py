from __future__ import annotations

import os
from dataclasses import dataclass
from pathlib import Path

from app.core.config import Settings
from app.services.text_utils import normalize_space, safe_read_text


TEXT_EXTENSIONS = {".md", ".txt", ".log", ".csv", ".json", ".yaml", ".yml", ".c", ".h", ".py"}
DOC_EXTENSIONS = TEXT_EXTENSIONS | {".pdf", ".docx", ".xlsx"}
SKIP_DIRS = {".git", "__pycache__", "node_modules", ".venv", ".venv310", ".venv312", ".next"}


@dataclass(frozen=True)
class DocumentChunk:
    chunk_id: str
    path: Path
    title: str
    text: str
    line_start: int | None = None
    line_end: int | None = None


class DocumentLoader:
    def __init__(self, settings: Settings):
        self.settings = settings
        self._chunk_cache: list[DocumentChunk] | None = None

    def iter_chunks(self, refresh: bool = False) -> list[DocumentChunk]:
        if self._chunk_cache is not None and not refresh:
            return self._chunk_cache
        chunks: list[DocumentChunk] = []
        for root in (self.settings.knowledge_root, self.settings.file_library_root):
            if not root.exists():
                continue
            chunks.extend(self._iter_root(root))
        self._chunk_cache = chunks
        return chunks

    def _iter_root(self, root: Path) -> list[DocumentChunk]:
        chunks: list[DocumentChunk] = []
        for dirpath, dirnames, filenames in os.walk(root):
            dirnames[:] = [name for name in dirnames if name not in SKIP_DIRS]
            for filename in filenames:
                path = Path(dirpath) / filename
                if path.suffix.lower() not in DOC_EXTENSIONS:
                    continue
                try:
                    if path.stat().st_size > self.settings.max_index_file_bytes:
                        continue
                    title = path.relative_to(root).as_posix()
                    chunks.extend(self._load_file(path, title))
                except Exception:
                    continue
        return chunks

    def _load_file(self, path: Path, title: str) -> list[DocumentChunk]:
        suffix = path.suffix.lower()
        if suffix in TEXT_EXTENSIONS:
            return self._chunk_text(path, title, safe_read_text(path))
        if suffix == ".pdf":
            return self._load_pdf(path, title)
        if suffix == ".docx":
            return self._load_docx(path, title)
        if suffix == ".xlsx":
            return self._load_xlsx(path, title)
        return []

    def _load_pdf(self, path: Path, title: str) -> list[DocumentChunk]:
        try:
            from pypdf import PdfReader
        except Exception:
            return []
        chunks: list[DocumentChunk] = []
        reader = PdfReader(str(path))
        for page_no, page in enumerate(reader.pages, 1):
            text = page.extract_text() or ""
            chunks.extend(self._chunk_text(path, f"{title}#page={page_no}", text, line_base=None))
        return chunks

    def _load_docx(self, path: Path, title: str) -> list[DocumentChunk]:
        try:
            from docx import Document
        except Exception:
            return []
        doc = Document(str(path))
        text = "\n".join(p.text for p in doc.paragraphs if p.text.strip())
        return self._chunk_text(path, title, text)

    def _load_xlsx(self, path: Path, title: str) -> list[DocumentChunk]:
        try:
            from openpyxl import load_workbook
        except Exception:
            return []
        wb = load_workbook(str(path), read_only=True, data_only=True)
        lines: list[str] = []
        for sheet in wb.worksheets:
            lines.append(f"[sheet] {sheet.title}")
            for row in sheet.iter_rows(values_only=True):
                cells = [normalize_space(str(cell)) for cell in row if cell is not None and str(cell).strip()]
                if cells:
                    lines.append(" | ".join(cells))
        return self._chunk_text(path, title, "\n".join(lines))

    def _chunk_text(
        self,
        path: Path,
        title: str,
        text: str,
        line_base: int | None = 1,
    ) -> list[DocumentChunk]:
        lines = [line.rstrip() for line in (text or "").splitlines()]
        chunks: list[DocumentChunk] = []
        current: list[str] = []
        start_line: int | None = line_base
        char_count = 0
        chunk_chars = max(300, self.settings.rag_chunk_chars)
        overlap = max(0, self.settings.rag_chunk_overlap)

        def flush(end_line: int | None) -> None:
            nonlocal current, start_line, char_count
            chunk_text = "\n".join(current).strip()
            if chunk_text:
                idx = len(chunks)
                safe_path = str(path).replace("\\", "/")
                chunks.append(
                    DocumentChunk(
                        chunk_id=f"{safe_path}::{title}::{idx}",
                        path=path,
                        title=title,
                        text=chunk_text[: chunk_chars + overlap],
                        line_start=start_line,
                        line_end=end_line,
                    )
                )
            if overlap and chunk_text:
                tail = chunk_text[-overlap:]
                current = [tail]
                char_count = len(tail)
                start_line = end_line
            else:
                current = []
                char_count = 0
                start_line = None

        for idx, line in enumerate(lines, 1):
            clean = normalize_space(line)
            if not clean:
                continue
            if start_line is None:
                start_line = idx if line_base is not None else None
            current.append(clean)
            char_count += len(clean) + 1
            if char_count >= chunk_chars:
                flush(idx if line_base is not None else None)

        if current:
            flush(len(lines) if line_base is not None else None)
        return chunks
