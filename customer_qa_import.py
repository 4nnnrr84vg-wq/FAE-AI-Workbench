"""从「客户问题汇总」Excel 导入问答到关键词库 learned.yaml。"""
from __future__ import annotations

import datetime as dt
import hashlib
import os
import re
import shutil
from pathlib import Path

import openpyxl
import yaml

from keyword_learning import KeywordLearner, extract_keywords
from text_encoding import write_text

_BASE = Path(__file__).resolve().parent

Q_HEADERS = frozenset({"问题", "问题描述", "客户问题"})
A_HEADERS = frozenset({"解决方法", "解决", "解决方案", "答复", "回答", "处理"})
STATUS_WORDS = frozenset({"已经解决", "已解决", "未解决", "进行中", "待处理", "已关闭"})
SKIP_PREFIXES = ("http://", "https://", "E:\\", "D:\\", "E:/", "D:/", "ftp://")


def customer_sheet_names(xlsx_path: str | os.PathLike | None = None) -> set[str]:
    """Excel 工作表名（客户名），不应出现在 keywords / 对外回复中。"""
    names: set[str] = set()
    path = xlsx_path or (_BASE / "keyword_data" / "imports" / "客户问题汇总.xlsx")
    p = Path(path)
    if not p.is_file():
        alt = Path(r"c:\Users\Administrator\Desktop\客户问题汇总(已自动还原).xlsx")
        if alt.is_file():
            p = alt
    if p.is_file():
        try:
            wb = openpyxl.load_workbook(str(p), read_only=True, data_only=True)
            names.update(wb.sheetnames)
            wb.close()
        except Exception:
            pass
    return names


def _strip_customer_names(text: str, names: set[str]) -> str:
    if not text or not names:
        return text
    out = text
    for name in sorted(names, key=len, reverse=True):
        if not name:
            continue
        out = out.replace(name, "")
    return re.sub(r"\s{2,}", " ", out).strip()


def _filter_keywords(keywords: list[str], names: set[str]) -> list[str]:
    blocked = {n.lower() for n in names if n}
    out: list[str] = []
    seen: set[str] = set()
    for kw in keywords:
        k = (kw or "").strip()
        if not k or k.lower() in blocked:
            continue
        if k.lower() in seen:
            continue
        seen.add(k.lower())
        out.append(k)
    return out


def _norm(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _find_header(cells: list[str]) -> dict[str, int]:
    cols: dict[str, int] = {}
    for i, text in enumerate(cells):
        if not text:
            continue
        if text in Q_HEADERS:
            cols["q"] = i
        elif text in A_HEADERS or text == "解决方法":
            cols["a"] = i
        elif text in ("备注", "JIRA", "Ticket"):
            cols["note"] = i
        elif text in ("SDK版本", "版本"):
            cols["ver"] = i
    return cols


def _is_meta_row(text: str) -> bool:
    if not text:
        return True
    low = text.lower()
    if any(text.startswith(p) for p in SKIP_PREFIXES):
        return True
    if text in Q_HEADERS or text in ("SDK版本", "版本", "时间"):
        return True
    if low.startswith("sheet") or "gitlab" in low or "gitee.com" in low:
        return True
    return False


def parse_customer_xlsx(path: str | os.PathLike) -> list[dict]:
    """解析 Excel，返回 {sheet, question, answer, note, version} 列表。"""
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    items: list[dict] = []
    seen_q: set[str] = set()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        cols: dict[str, int] = {}

        for row in ws.iter_rows(values_only=True):
            cells = [_norm(c) for c in row]
            if not any(cells):
                continue

            hdr = _find_header(cells)
            if hdr.get("q") is not None:
                cols = hdr
                continue

            if "q" not in cols:
                continue

            qi = cols["q"]
            if qi >= len(cells):
                continue
            question = cells[qi]
            if _is_meta_row(question) or len(question) < 3:
                continue

            answer = ""
            if cols.get("a") is not None and cols["a"] < len(cells):
                answer = cells[cols["a"]]
            note = ""
            if cols.get("note") is not None and cols["note"] < len(cells):
                note = cells[cols["note"]]
            version = ""
            if cols.get("ver") is not None and cols["ver"] < len(cells):
                version = cells[cols["ver"]]

            if answer in STATUS_WORDS:
                answer = ""
            if not answer and note and note not in STATUS_WORDS:
                answer = note
                note = ""

            # 部分表把状态写在「备注」前一列
            for extra in cells:
                if extra in STATUS_WORDS:
                    continue

            if note and answer and note not in answer:
                if not note.startswith("KXYJ") and not note.startswith("LAMP"):
                    answer = f"{answer}（备注：{note}）"
                else:
                    answer = f"{answer}（{note}）"

            q_key = re.sub(r"\s+", " ", question.lower())[:200]
            if q_key in seen_q:
                continue
            seen_q.add(q_key)

            items.append(
                {
                    "sheet": sheet_name,
                    "question": question,
                    "answer": answer,
                    "note": note,
                    "version": version,
                }
            )

    wb.close()
    return items


def _slug_id(question: str) -> str:
    return "cust_" + hashlib.md5(question.encode("utf-8", errors="ignore")).hexdigest()[:10]


def import_customer_xlsx(
    cfg: dict,
    xlsx_path: str | os.PathLike,
    *,
    min_answer_len: int = 4,
    priority: int | None = None,
    dry_run: bool = False,
) -> tuple[int, int, str]:
    """
    导入客户问题 Excel 到 learned.yaml。
    返回 (成功条数, 跳过条数, 说明信息)
    """
    path = Path(xlsx_path)
    if not path.is_file():
        return 0, 0, f"文件不存在: {path}"

    cq = cfg.get("customer_qa", {}) or {}
    pri = int(priority if priority is not None else cq.get("priority", 88))
    min_ans = int(cq.get("min_answer_len", min_answer_len))
    max_reply = int(cq.get("max_reply_len", 1200))

    parsed = parse_customer_xlsx(path)
    learner = KeywordLearner(cfg)
    sheet_names = customer_sheet_names(path)
    added = 0
    skipped = 0
    now = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    for item in parsed:
        q = item["question"].strip()
        a = item["answer"].strip()
        if len(a) < min_ans:
            skipped += 1
            continue

        if len(a) > max_reply:
            a = a[:max_reply] + "\n…（已截断）"
        a = _strip_customer_names(a, sheet_names)

        keywords = extract_keywords(q)
        if not keywords and len(q) >= 3:
            keywords = [q[:48]]
        keywords = _filter_keywords(keywords, sheet_names)
        if not keywords:
            skipped += 1
            continue

        for chip in re.findall(r"\b(XC\d{3,5}[A-Z0-9]*|2421-\d+|6012)\b", q, re.I):
            if chip not in keywords:
                keywords.append(chip)

        entry_id = _slug_id(q)
        new_entry = {
            "id": entry_id,
            "source": "customer_xlsx",
            "priority": pri,
            "keywords": keywords[:12],
            "match": "contains",
            "reply": a,
            "question": q[:200],
            "created": now,
            "hits": 0,
        }

        replaced = False
        for i, e in enumerate(learner._entries):
            if e.get("id") == entry_id:
                new_entry["hits"] = int(e.get("hits", 0))
                new_entry["created"] = e.get("created", now)
                learner._entries[i] = new_entry
                replaced = True
                break
        if not replaced:
            # 同问题不同 id：合并关键词
            qn = q.lower()[:80]
            merged = False
            for e in learner._entries:
                if str(e.get("source", "")) != "customer_xlsx":
                    continue
                if str(e.get("question", "")).lower()[:80] == qn:
                    kws = list(e.get("keywords") or [])
                    for kw in keywords:
                        if kw not in kws:
                            kws.append(kw)
                    e["keywords"] = _filter_keywords(kws, sheet_names)[:12]
                    e["reply"] = a
                    e.pop("customer", None)
                    merged = True
                    break
            if not merged:
                learner._entries.append(new_entry)
        added += 1

    if dry_run:
        return added, skipped, f"[预览] 可导入 {added} 条，跳过 {skipped} 条（无有效答案）"

    if added:
        if len(learner._entries) > learner.max_entries:
            learner._entries = learner._entries[-learner.max_entries :]
        learner._save()

    dest = _BASE / "keyword_data" / "imports" / path.name
    if path.resolve() != dest.resolve():
        dest.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(path, dest)

    return (
        added,
        skipped,
        f"客户问题导入完成：{added} 条写入 learned.yaml，跳过 {skipped} 条（无答案过短）。"
        f"\n源文件已复制到 {dest}",
    )


def customer_qa_status(cfg: dict) -> str:
    cq = cfg.get("customer_qa", {}) or {}
    xlsx = cq.get("xlsx_file", "keyword_data/imports/客户问题汇总.xlsx")
    p = Path(xlsx)
    if not p.is_absolute():
        p = _BASE / p
    learner = KeywordLearner(cfg)
    n = sum(1 for e in learner.entries() if str(e.get("source")) == "customer_xlsx")
    lines = [
        f"客户问题库：learned.yaml 中 customer_xlsx 条目 {n} 条",
        f"Excel：{p} {'(存在)' if p.is_file() else '(未找到)'}",
        "导入命令：/keyword 导入客户问题  或运行 scripts/import_customer_qa.py",
    ]
    return "\n".join(lines)


def scrub_learned_customer_names(cfg: dict, xlsx_path: str | os.PathLike | None = None) -> tuple[int, str]:
    """清理 learned.yaml 中 customer_xlsx 条目的客户名 keywords 与 customer 字段。"""
    names = customer_sheet_names(xlsx_path)
    learner = KeywordLearner(cfg)
    n = 0
    for e in learner._entries:
        if str(e.get("source", "")) != "customer_xlsx":
            continue
        if e.pop("customer", None) is not None:
            n += 1
        old_kws = list(e.get("keywords") or [])
        new_kws = _filter_keywords(old_kws, names)
        if new_kws != old_kws:
            e["keywords"] = new_kws[:12]
            n += 1
        old_reply = str(e.get("reply", ""))
        new_reply = _strip_customer_names(old_reply, names)
        if new_reply != old_reply:
            e["reply"] = new_reply
            n += 1
    if n:
        learner._save()
    return n, f"已清理 {n} 处客户名引用（keywords / customer 字段 / reply）"
